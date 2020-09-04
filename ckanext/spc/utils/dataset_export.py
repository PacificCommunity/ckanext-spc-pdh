import json
import logging
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import ckan.plugins.toolkit as tk

log = logging.getLogger(__name__)


def _make_export_doc():
    doc = Workbook()
    del doc["Sheet"]
    return doc


def _simplified_schemas():

    schemas = {
        name: {
            "dataset": _simplify_fields(options["dataset_fields"]),
            "resource": _simplify_fields(options["resource_fields"])
            + [{"name": "url_type"}],
        }
        for name, options in tk.h.scheming_dataset_schemas().items()
    }
    return schemas


def _simplify_fields(fields):
    return [
        {"name": f["field_name"]}
        for f in sorted(fields, key=lambda d: d["field_name"])
    ]


def _write_into_doc(pkg, doc: Workbook, schema):
    type_ = pkg["type"]
    res_sheet_name = f"{type_} resources"
    if type_ not in doc:
        doc.create_sheet(type_)
        doc.create_sheet(res_sheet_name)

    _fill_sheet(pkg, doc[type_], schema["dataset"])
    for res in pkg["resources"]:
        _fill_sheet(res, doc[res_sheet_name], schema["resource"])


def _fill_sheet(data, sheet: Worksheet, fields):
    if sheet.max_row == 1:
        for col, field in enumerate(fields, 1):
            sheet.cell(1, col, field["name"])
    row = sheet.max_row + 1
    for col, field in enumerate(fields, 1):
        value = data.get(field["name"])
        if isinstance(value, (list, dict)):
            value = json.dumps(value)
        sheet.cell(row, col, value)


class Exporter(object):
    def __init__(self, collection=[]):
        self.doc = _make_export_doc()
        self.schemas = _simplified_schemas()
        self.collection = collection

    def _generate(self):
        for pkg in self.collection:
            try:
                schema = self.schemas[pkg["type"]]
            except KeyError as e:
                log.error("Cannot define package schema: %s", e)
                continue
            _write_into_doc(pkg, self.doc, schema)

    def get_stream(self):
        self._generate()
        with NamedTemporaryFile(mode="w+b") as temp:
            self.doc.save(temp.name)
            self.doc.save("/tmp/x.xlsx")
            temp.seek(0)
            yield temp.read()
