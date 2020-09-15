import json
import logging

from tempfile import NamedTemporaryFile
from typing import Any, Dict, List

import funcy as F
from funcy.funcs import identity

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import ckan.plugins.toolkit as tk

log = logging.getLogger(__name__)

safe_loads = F.ignore((ValueError, TypeError))(json.loads)


def _make_export_doc():
    doc = Workbook()
    del doc["Sheet"]
    return doc


def _simplified_schemas() -> Dict[str, List[Dict[str, Any]]]:

    schemas = {
        name: {
            "dataset": _simplify_fields(options["dataset_fields"])
            + [{"name": "id"}, {"name": "type"}],
            "resource": _simplify_fields(options["resource_fields"])
            + [{"name": "id"}, {"name": "url_type"}, {"name": "package_id"}],
        }
        for name, options in tk.h.scheming_dataset_schemas().items()
    }
    return schemas


def _simplify_fields(fields):
    return [
        {"name": f["field_name"]}
        for f in sorted(fields, key=lambda d: d["field_name"])
    ]


def _resource_sheet_name(type_: str):
    return f"{type_} resources"


def _write_into_doc(pkg, doc: Workbook, schema):
    type_ = pkg["type"]
    res_sheet_name = _resource_sheet_name(type_)
    if type_ not in doc:
        doc.create_sheet(type_)
        doc.create_sheet(res_sheet_name)

    _fill_sheet(pkg, doc[type_], schema["dataset"])
    for res in pkg["resources"]:
        _fill_sheet(res, doc[res_sheet_name], schema["resource"])


def _fill_sheet(data, sheet: Worksheet, fields):
    if sheet.max_row == 1:
        for col, field in enumerate(fields, 1):
            sheet.cell(1, col, field["name"])
    row = sheet.max_row + 1
    for col, field in enumerate(fields, 1):
        value = data.get(field["name"])
        if isinstance(value, (list, dict)):
            value = json.dumps(value)
        sheet.cell(row, col, value)


def _read_sheet(sheet: Worksheet):
    rows = sheet.rows
    fields = [cell.value for cell in next(rows)]
    return [
        dict(
            zip(
                fields,
                [
                    F.iffy(F.notnone, F.identity, identity(cell.value))(
                        safe_loads(
                            (cell.value or "")
                            .replace("‘", "'")
                            .replace("’", "'")
                            .replace("´", "'")
                            .replace("“", '"')
                            .replace("”", '"')
                        )
                    )
                    for cell in row
                ],
            )
        )
        for row in rows
    ]


def _attach_resources(resources, datasets):
    for resource in resources:
        try:
            dataset = datasets[resource["package_id"]]
        except KeyError:
            log.error(
                "Can't locate package %s for resource %s",
                resource["package_id"],
                resource["id"],
            )
            continue
        dataset.setdefault("resources", []).append(resource)


class Exporter(object):
    def __init__(self, collection=[]):
        self.doc = _make_export_doc()
        self.schemas = _simplified_schemas()
        self.collection = collection

    def _generate(self):
        for pkg in self.collection:
            try:
                schema = self.schemas[pkg["type"]]
            except KeyError as e:
                log.error("Cannot define package schema: %s", e)
                continue
            _write_into_doc(pkg, self.doc, schema)

    def get_stream(self):
        self._generate()
        with NamedTemporaryFile(mode="w+b") as temp:
            self.doc.save(temp.name)
            self.doc.save("/tmp/x.xlsx")
            temp.seek(0)
            yield temp.read()


class Importer(object):
    def __init__(self, stream):
        self.doc = load_workbook(stream, read_only=True)
        self.schemas = _simplified_schemas()
        self.collection = []

    def _generate(self):
        for sheet in self.doc.worksheets:
            if sheet.title in self.schemas:
                datasets = _read_sheet(sheet)
                resources = _read_sheet(
                    self.doc[_resource_sheet_name(sheet.title)]
                )
                _attach_resources(resources, {d["id"]: d for d in datasets})
                self.collection.extend(datasets)

    def get_datasets(self):
        self._generate()
        return self.collection
